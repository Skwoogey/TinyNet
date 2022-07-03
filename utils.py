import tensorflow as tf
import tensorflow_addons as tfa
from depthwise3DConv import DepthwiseConv3D
from CenterLossLayer import CenterLossLayer

import numpy as np
import math
import scipy.io as scp
from network import makeFast3DCNN
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook
import random

def saveTrainInfo(history, model_for_save, path):
	metrics_keys = list(history.history.keys())
	num_of_epochs = len(list(history.history.values())[0])
	print(num_of_epochs)

	wb = Workbook()
	ws = wb.active

	for metric_i in range(len(metrics_keys)):
		ws.cell(row = 1, column = metric_i + 1).value = metrics_keys[metric_i]

	for epoch_i in range(num_of_epochs):
		for metric_i in range(len(metrics_keys)):
			ws.cell(row = epoch_i + 2, column = metric_i + 1).value = history.history[metrics_keys[metric_i]][epoch_i]

	layer_i = 0
	while True:
		try:
			layer = model_for_save.get_layer(index=layer_i)
			ws.cell(row = layer_i + 1, column = len(metrics_keys) + 3).value = layer.name
			ws.cell(row = layer_i + 1, column = len(metrics_keys) + 4).value = str(type(layer))
			ws.cell(row = layer_i + 1, column = len(metrics_keys) + 5).value = str(layer.output.shape)
			ws.cell(row = layer_i + 1, column = len(metrics_keys) + 6).value = layer.count_params()
			layer_i += 1
		except ValueError:
			#print("value error", layer_i)
			break

	wb.save(filename = path)

def printEvaluation(y_pred, y_true):
	print("Calculating metrics...")
	num_classes = len(np.unique(y_true))
	confusion_matrix = np.zeros((num_classes, num_classes), dtype = int)

	labels_zip = np.stack([y_true, y_pred], axis = 1)

	for label_pair in labels_zip:
		#print(label_pair)
		confusion_matrix[label_pair[1], label_pair[0]] += 1
	

	sum_horizontal = np.sum(confusion_matrix, axis = 1)
	sum_vertical = np.sum(confusion_matrix, axis = 0)
	diagonal = np.diagonal(confusion_matrix)
	total = np.sum(confusion_matrix)
	p_c = np.sum(sum_horizontal * sum_vertical) / total
	p_0 = np.sum(diagonal)

	OA = np.trace(confusion_matrix) / total
	PA = diagonal / sum_vertical
	UA = diagonal / sum_horizontal
	K = (p_0 - p_c) / (total - p_c)
	print("Confusion matrix:")
	print(confusion_matrix)
	
	print("Total labels:", total)
	print('Overall accuracy:', OA)
	print('Producer\'s accuracy:', PA)
	print('User\'s accuracy:', UA)
	print('Kappa accuracy:', K)



def loadData(data_fl, gt_fl):
	data = scp.loadmat(data_fl)['data']
	gt = scp.loadmat(gt_fl)['data']

	return data, gt

def padImage(image, window_size, mode = 'edge'):
	pad_on_side =  (window_size - 1) // 2
	print(pad_on_side)
	padded_image = np.pad(
		image, 
		(
			(pad_on_side, pad_on_side),
			(pad_on_side, pad_on_side),
			(0, 0)
		),
		mode
	)

	return padded_image

def createTrainDataPerClass(image, gt, window_size, return_positions=False):
	num_of_classified_pixels = np.count_nonzero(gt)
	bands = image.shape[2]

	_, counts = np.unique(gt, return_counts=True)
	counts = counts[1:]
	num_of_classes = len(counts)

	data_per_class = []
	pixel_positions = []
	for class_i in range(num_of_classes):
		data_per_class.append(
			np.ndarray((counts[class_i], window_size, window_size, bands))
		)
		pixel_positions.append([])
		print(class_i + 1, counts[class_i], data_per_class[-1].shape, data_per_class[-1].dtype)

	original_image_size = (
		image.shape[0] - window_size + 1,
		image.shape[1] - window_size + 1,
	)
	window_radius = (window_size - 1) // 2

	i = [0] * num_of_classes
	for x in range(original_image_size[0]):
		for y in range(original_image_size[1]):
			if gt[x, y] != 0:
				label = gt[x, y] - 1
				pixel_positions[label].append((x, y))
				data_per_class[label][i[label]] = image[x:x + window_size, y:y + window_size, :]
				#assert((data_per_class[label][i[label], window_radius, window_radius, :] == image[x + window_radius, y + window_radius, :]).all())
				i[label] += 1
	if return_positions:
		return data_per_class, pixel_positions
	return data_per_class

def createTrainValidTestSplits(data, train_ratio, valid_ratio, test_ratio):
	normalization_sum = train_ratio + valid_ratio + test_ratio
	train_ratio /= normalization_sum
	valid_ratio /= normalization_sum
	test_ratio /= normalization_sum
	data_shape = (data[0].shape[1], data[0].shape[2], data[0].shape[3])

	num_of_cubes_per_class = np.ndarray((len(data), 3), dtype=int)

	for class_i in range(len(data)):
		num_of_pixels = data[class_i].shape[0]
		num_of_cubes_per_class[class_i, 0] = math.ceil(num_of_pixels * train_ratio)
		num_of_cubes_per_class[class_i, 1] = math.ceil(num_of_pixels * valid_ratio)
		num_of_cubes_per_class[class_i, 2] = num_of_pixels -\
			num_of_cubes_per_class[class_i, 1] - \
			num_of_cubes_per_class[class_i, 0]

	#print(num_of_cubes_per_class)
	pixels_per_split = np.sum(num_of_cubes_per_class, axis = 0)
	print("pixels per split", pixels_per_split)

	train_data = np.ndarray((pixels_per_split[0], *data_shape), dtype=np.float64)
	train_labels = np.ndarray((pixels_per_split[0], ), dtype=int)
	print("Train data shape", train_data.shape)

	valid_data = None
	valid_labels = None
	if valid_ratio != 0.0:
		valid_data = np.ndarray((pixels_per_split[1], *data_shape), dtype=np.float64)
		valid_labels = np.ndarray((pixels_per_split[1], ), dtype=int)
		print("Valid data shape", valid_data.shape)

	test_data = np.ndarray((pixels_per_split[2], *data_shape), dtype=np.float64)
	test_labels = np.ndarray((pixels_per_split[2], ), dtype=int)
	print("Test data shape", test_data.shape)

	tvt_index = np.zeros((3,), dtype=int)

	for class_i in range(len(data)):
		cur_class = data[class_i]
		indecies = list(range(cur_class.shape[0]))
		cur_class_size = num_of_cubes_per_class[class_i]
		random.shuffle(indecies)

		print("Class", class_i)
		print("split numbers", cur_class_size)

		start_index = 0
		train_indecies = indecies[start_index: start_index + cur_class_size[0]]
		start_index   += cur_class_size[0]
		valid_indecies = indecies[start_index: start_index + cur_class_size[1]]
		start_index   += cur_class_size[1]
		test_indecies  = indecies[start_index: start_index + cur_class_size[2]]

		train_data[tvt_index[0]: tvt_index[0] + cur_class_size[0]] = cur_class[train_indecies]
		train_labels[tvt_index[0]: tvt_index[0] + cur_class_size[0]] = class_i

		if valid_ratio != 0.0:
			valid_data[tvt_index[1]: tvt_index[1] + cur_class_size[1]] = cur_class[valid_indecies]
			valid_labels[tvt_index[1]: tvt_index[1] + cur_class_size[1]] = class_i

		test_data[tvt_index[2]: tvt_index[2] + cur_class_size[2]] = cur_class[test_indecies]
		test_labels[tvt_index[2]: tvt_index[2] + cur_class_size[2]] = class_i

		tvt_index = tvt_index + cur_class_size

	return  [train_data, train_labels], [valid_data, valid_labels], [test_data, test_labels]

def createCustomSplits(data, ratios):
	normalization_sum = sum(ratios)
	ratios =  [x / normalization_sum for x in ratios]
	data_shape = (data[0].shape[1], data[0].shape[2], data[0].shape[3])

	num_of_cubes_per_class = np.zeros((len(data), len(ratios)), dtype=int)

	for class_i in range(len(data)):
		num_of_pixels = data[class_i].shape[0]

		for split_i in range(len(ratios) - 1):
			num_of_cubes_per_class[class_i, split_i] = max(1, math.ceil(round(num_of_pixels * ratios[split_i])))

		num_of_cubes_per_class[class_i, -1] = num_of_pixels - np.sum(num_of_cubes_per_class[class_i, :-1])

	#print(num_of_cubes_per_class)
	pixels_per_split = np.sum(num_of_cubes_per_class, axis = 0)
	print("pixels per split", pixels_per_split)

	splits_data = []
	splits_labels = []
	for split_i in range(len(ratios)):
		splits_data.append(np.ndarray((pixels_per_split[split_i], *data_shape), dtype=np.float64))
		splits_labels.append(np.ndarray((pixels_per_split[split_i], ), dtype=int))

	tvt_index = np.zeros((len(ratios),), dtype=int)

	for class_i in range(len(data)):
		cur_class = data[class_i]
		indecies = list(range(cur_class.shape[0]))
		cur_class_size = num_of_cubes_per_class[class_i]
		random.shuffle(indecies)

		print("Class", class_i)
		print('Total pixels:', data[class_i].shape[0])
		print("split numbers", cur_class_size)

		start_index = 0

		for split_i in range(len(ratios)):
			split_indecies = indecies[start_index: start_index + cur_class_size[split_i]]
			splits_data[split_i][tvt_index[split_i]: tvt_index[split_i] + cur_class_size[split_i]] = cur_class[split_indecies]
			splits_labels[split_i][tvt_index[split_i]: tvt_index[split_i] + cur_class_size[split_i]] = class_i

			start_index += cur_class_size[split_i]

		tvt_index = tvt_index + cur_class_size

	return  splits_data, splits_labels

def getLayersFromModel(model):
	layers = []
	layer_i = 0
	while True:
		try:
			layer = model.get_layer(index=layer_i)
			layers.append(layer)
			layer_i += 1
		except ValueError:
			break

	return layers

def copyPretrainedModel(dst_model, src_model_path):
	custom_objects = {
		'DepthwiseConv3D': DepthwiseConv3D,
		'AdamW': tfa.optimizers.AdamW,
		'SGDW' : tfa.optimizers.SGDW,
		'CenterLossLayer': CenterLossLayer
	}
	pretrained_model = tf.keras.models.load_model(src_model_path, custom_objects=custom_objects)

	dest_layers = getLayersFromModel(dst_model)
	for dl in dest_layers:
		if len(dl.get_weights()) == 0:
			dest_layers.remove(dl)
	src_layers = getLayersFromModel(pretrained_model)
	for sl in src_layers:
		if len(sl.get_weights()) == 0:
			src_layers.remove(sl)

	if len(src_layers) != len(dest_layers):
		raise ValueError('incompatible models: different amount of layers')

	for sl, dl in zip(src_layers[:-1], dest_layers[:-1]):
		dl.set_weights(sl.get_weights())
		for slw, dlw in zip(sl.get_weights(), dl.get_weights()):
			assert((slw == dlw).all())

	src_last = src_layers[-1]
	dst_last = dest_layers[-1]
	sw = src_last.get_weights()
	dw = dst_last.get_weights()
	class_dif = dw[0].shape[1] - sw[0].shape[1]
	dw[0].shape[1]
	if class_dif == 0:
		pass
	elif class_dif < 0:
		sw[0] = sw[0][:, :dw[0].shape[1]]
		sw[1] = sw[1][:dw[0].shape[1]]
	else:
		sw[0] = np.concatenate([sw[0]]*3, axis = 1)
		sw[1] = np.concatenate([sw[1]]*3, axis = 1)
		sw[0] = sw[0][:, :dw[0].shape[1]]
		sw[1] = sw[1][:dw[0].shape[1]]

	dst_last.set_weights(sw)